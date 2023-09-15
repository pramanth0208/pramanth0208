10.a) Write a python program to combine select pages from many PDFs from PyPDF2 

import PdfWriter, PdfReader

from PyPDF2 import PdfWriter, PdfReader

num = int(input("Enter page number you want to combine from multiple documents "))

pdf1 = open('8a.pdf', 'rb')

pdf2 = open('9a&9b.pdf', 'rb')

pdf_writer = PdfWriter()

pdf1_reader = PdfReader(pdf1)

page = pdf1_reader.pages[num - 1]

pdf_writer.add_page(page)

pdf2_reader = PdfReader(pdf2)

page = pdf2_reader.pages[num - 1]

pdf_writer.add_page(page)

with open('output.pdf', 'wb') as output:

 pdf_writer.write(output)

10.b) Write a python program to fetch current weather data from the JSON file

import requests

def fetch_weather_data(api_key, city):

 base_url = "http://api.openweathermap.org/data/2.5/weather"

 params = {

 "q": city,

 "appid": api_key,

 "units": "metric" # You can change this to "imperial" for Fahrenheit

 }

 response = requests.get(base_url, params=params)

 weather_data = response.json()

 return weather_data

def display_weather_data(weather_data):

 print("Weather in", weather_data["name"])

 print("Temperature:", weather_data["main"]["temp"], "°C")

 print("Description:", weather_data["weather"][0]["description"])

if __name__ == "__main__":

 # Replace with your OpenWeatherMap API key

 api_key = "Google Maps API"

 city = input("Enter city name: ")

 weather_data = fetch_weather_data(api_key, city)

 if "cod" in weather_data and weather_data["cod"] == 200:

 display_weather_data(weather_data)

 else:

 print("City not found or error fetching weather data.")


9.b) Demonstrate python program to read the data from the spreadsheet and write the 

data in to the spreadsheet

from openpyxl import Workbook

from openpyxl.styles import Font

wb = Workbook()

sheet = wb.active

sheet.title = "Language"

wb.create_sheet(title = "Capital")

lang = ["Kannada", "Telugu", "Tamil"]

state = ["Karnataka", "Telangana", "Tamil Nadu"]

capital = ["Bengaluru", "Hyderabad", "Chennai"]

code =['KA', 'TS', 'TN']

sheet.cell(row = 1, column = 1).value = "State"

sheet.cell(row = 1, column = 2).value = "Language"

sheet.cell(row = 1, column = 3).value = "Code"

ft = Font(bold=True)

for row in sheet["A1:C1"]:

 for cell in row:

 cell.font = ft

for i in range(2,5):

 sheet.cell(row = i, column = 1).value = state[i-2]

 sheet.cell(row = i, column = 2).value = lang[i-2]

 sheet.cell(row = i, column = 3).value = code[i-2]

wb.save("demo.xlsx")

sheet = wb["Capital"]

sheet.cell(row = 1, column = 1).value = "State"

sheet.cell(row = 1, column = 2).value = "Capital"

sheet.cell(row = 1, column = 3).value = "Code"

ft = Font(bold=True)

for row in sheet["A1:C1"]:

 for cell in row:

 cell.font = ft

for i in range(2,5):

 sheet.cell(row = i, column = 1).value = state[i-2]

 sheet.cell(row = i, column = 2).value = capital[i-2]

 sheet.cell(row = i, column = 3).value = code[i-2]

wb.save("demo.xlsx")
